import openpyxl as xl


wb = xl.load_workbook('transactions.xlsx')
#In wb['Sheet'] sheet is case sensitive an needs an uppercase letter at the start
sheet = wb['Sheet1']
#cell = sheet.cell(1, 1)
#print(cell.value)
#max_row = sheet.max_row
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('transactons2.xlsx')
